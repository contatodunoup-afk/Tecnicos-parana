import openpyxl
import json

# Carregar o arquivo Excel
wb = openpyxl.load_workbook('dados fundepar tecnicos.xlsx')
ws = wb.active

print("Sheets disponíveis:", wb.sheetnames)
print("\nDados da planilha (primeiras 25 linhas):")
for i, row in enumerate(ws.iter_rows(max_row=25, values_only=True), 1):
    print(f"Linha {i}: {row}")
