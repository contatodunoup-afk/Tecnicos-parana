import openpyxl
import json

wb = openpyxl.load_workbook('dados fundepar tecnicos.xlsx')
ws = wb.active

empresas = []
# Rows começam em 2 (row 1 é header)
for row in range(2, ws.max_row + 1):
    categoria = ws.cell(row, 2).value
    name = ws.cell(row, 4).value
    phone = ws.cell(row, 5).value
    code = ws.cell(row, 6).value
    city = ws.cell(row, 15).value
    state = ws.cell(row, 14).value
    lat = ws.cell(row, 29).value
    lng = ws.cell(row, 30).value
    
    # Verificar se tem dados essenciais
    if name and lat and lng:
        # Remover RODRIACO (code 068178)
        if str(code).strip() != '068178':
            empresas.append({
                'name': str(name).strip(),
                'city': str(city).strip() if city else '',
                'state': str(state).strip() if state else '',
                'phone': str(phone).strip() if phone else '',
                'code': str(code).strip() if code else '',
                'lat': float(lat),
                'lng': float(lng),
                'classificacao': str(categoria).strip() if categoria else 'Padrao'
            })

print(f'Total de registros: {len(empresas)}')
for e in empresas:
    print(f"  {e['name']} - Classificação: {e['classificacao']}")

# Salvar em JSON
with open('data_fundepar.json', 'w', encoding='utf-8') as f:
    json.dump(empresas, f, ensure_ascii=False, indent=2)

print(f'\nArquivo salvo: data_fundepar.json')
